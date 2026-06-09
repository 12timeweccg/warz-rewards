import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl


WORKSPACE = Path(r"C:\Users\almon\OneDrive\Desktop\WARZ WEBSITE PUBLISHED")
OUTPUT = WORKSPACE / "events-data.js"
SKIP_SHEETS = {"Data", "Template", "รายชื่อกิจกรรม"}
EXCLUDED_EVENTS = {"ภารกิจค้นหาเสบียงลับ"}
ICONS = ["broadcast", "share", "calendar", "clock"]


def resolve_source():
    candidates = sorted(
        WORKSPACE.glob("*ประกาศรายชื่อกิจกรรม*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("ไม่พบไฟล์ประกาศรายชื่อกิจกรรมใน workspace")
    return candidates[0]


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def clean_name(name):
    cleaned = re.sub(r"[💛💜⚙️🎁🕹🌐✅📌🩵🧡]", "", name).strip()
    return cleaned or name.strip()


def normalize_text(value):
    return cell_text(value).strip().lower()


def find_header_row(sheet):
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [cell_text(value) for value in row]
        if "NO." in values and "Reward" in values and "Facebook" in values and "UID" in values:
            return row_index
    return None


def build_item_map(workbook):
    if "Data" not in workbook.sheetnames:
        return {}

    sheet = workbook["Data"]
    item_map = {}
    for row in sheet.iter_rows(min_row=4, values_only=True):
        values = [cell_text(value) for value in row]
        item_id = values[2] if len(values) > 2 else ""
        image_url = values[3] if len(values) > 3 else ""
        item_name = values[4] if len(values) > 4 else ""
        item_type = values[8] if len(values) > 8 else ""
        if item_id:
            item_map[item_id] = {
                "imageUrl": image_url,
                "name": item_name,
                "type": item_type,
            }
    return item_map


_HEADER_LITERALS = {"item id", "item en", "item type", "own period", "amount", "image url", "image", "bundle"}


def _is_header_val(val):
    return val.strip().lower() in _HEADER_LITERALS


def build_reward_info(values, item_map):
    reward = values[1] if len(values) > 1 else ""
    bundle = values[8] if len(values) > 8 else ""
    item_id = values[10] if len(values) > 10 else ""
    item_en = values[11] if len(values) > 11 else ""
    own_period = values[12] if len(values) > 12 else ""
    item_type = values[13] if len(values) > 13 else ""
    amount = values[14] if len(values) > 14 else ""
    image_url = values[16] if len(values) > 16 else ""

    # Skip rows where item columns contain Excel header labels
    if _is_header_val(item_id) or _is_header_val(item_en) or _is_header_val(amount) or _is_header_val(image_url):
        item_id = ""
        item_en = ""
        item_type = ""
        own_period = ""
        amount = ""
        image_url = ""

    item_data = item_map.get(item_id, {})
    display_item_en = item_en or item_data.get("name", "")

    if not any([reward, bundle, item_id, display_item_en, image_url]):
        return None

    return {
        "name": reward or "รางวัลกิจกรรม",
        "forumReward": reward or "รางวัลกิจกรรม",
        "itemEn": display_item_en,
        "itemId": item_id,
        "itemType": item_type or item_data.get("type", ""),
        "ownPeriod": own_period,
        "amount": amount or "1",
        "imageUrl": image_url or item_data.get("imageUrl", ""),
        "hasItem": bool(display_item_en or item_id or image_url or item_data.get("imageUrl", "")),
    }


def is_item_header_row(values):
    lowered = [normalize_text(value) for value in values]
    return "item id" in lowered and "item en" in lowered


def build_header_map(values):
    return {normalize_text(value): index for index, value in enumerate(values) if normalize_text(value)}


def value_from_map(values, header_map, key):
    index = header_map.get(key)
    if index is None or index >= len(values):
        return ""
    return values[index]


def build_block_reward_info(values, header_map, item_map):
    item_id = value_from_map(values, header_map, "item id")
    item_en = value_from_map(values, header_map, "item en")
    amount = value_from_map(values, header_map, "amount")
    image_url = value_from_map(values, header_map, "image url")
    item_type = value_from_map(values, header_map, "item type")
    own_period = value_from_map(values, header_map, "own period")

    item_data = item_map.get(item_id, {})
    display_item_en = item_en or item_data.get("name", "")
    resolved_image = image_url or item_data.get("imageUrl", "")

    if not any([item_id, display_item_en, resolved_image]):
        return None

    return {
        "itemEn": display_item_en,
        "itemId": item_id,
        "amount": amount or "1",
        "imageUrl": resolved_image,
        "itemType": item_type or item_data.get("type", ""),
        "ownPeriod": own_period,
    }


def extract_code(value):
    text = cell_text(value)
    match = re.search(r"(?:master\s*code|item\s*code)\s*:\s*([A-Z0-9-]+)", text, re.I)
    if match:
        return match.group(1).strip()
    return ""


def unique_rewards(rewards):
    seen = set()
    results = []
    for reward in rewards:
        key = (reward.get("itemId") or reward.get("itemEn") or reward.get("imageUrl") or reward.get("name")).strip()
        if not key or key in seen:
            continue
        seen.add(key)
        results.append(reward)
    return results


def find_master_codes(sheet, title, work_status, claim_end, item_map):
    item_blocks = []
    code_rows = []
    current_block = None
    inside_items = False
    current_header_map = None

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [cell_text(value) for value in row]

        for value in values:
            code = extract_code(value)
            if code:
                code_rows.append({"row": row_index, "code": code})

        if is_item_header_row(values):
            if current_block and current_block["items"]:
                item_blocks.append(current_block)
            current_block = {"start": row_index, "items": []}
            inside_items = True
            current_header_map = build_header_map(values)
            continue

        if not inside_items:
            continue

        if current_block is None:
            current_block = {"start": row_index, "items": []}

        reward_info = build_block_reward_info(values, current_header_map or {}, item_map)
        if reward_info:
            current_block["items"].append(
                {
                    "itemEn": reward_info["itemEn"],
                    "itemId": reward_info["itemId"],
                    "amount": reward_info["amount"],
                    "imageUrl": reward_info["imageUrl"],
                }
            )
            current_block["end"] = row_index
            continue

        first_cell = normalize_text(values[0] if values else "")
        if first_cell == "รวม" or first_cell == "no.":
            if current_block["items"]:
                item_blocks.append(current_block)
            current_block = None
            inside_items = False
            current_header_map = None

    if current_block and current_block["items"]:
        item_blocks.append(current_block)

    master_codes = []
    for code_row in code_rows:
        associated_block = None
        previous_blocks = [block for block in item_blocks if block.get("end", block["start"]) < code_row["row"]]
        next_blocks = [block for block in item_blocks if block["start"] > code_row["row"]]

        if previous_blocks:
            associated_block = previous_blocks[-1]
        elif next_blocks:
            associated_block = next_blocks[0]

        master_codes.append(
            {
                "code": code_row["code"],
                "eventName": title,
                "status": work_status or "พร้อมใช้",
                "expiresAt": claim_end or "-",
                "items": unique_rewards((associated_block or {}).get("items", [])),
            }
        )

    return master_codes


def build_events(workbook):
    item_map = build_item_map(workbook)
    events = []
    master_codes = []

    for sheet in workbook.worksheets:
        if sheet.title in SKIP_SHEETS:
            continue

        title = clean_name(sheet.title)
        if title in EXCLUDED_EVENTS:
            continue

        delivery_date = cell_text(sheet.cell(2, 2).value)
        work_status = cell_text(sheet.cell(3, 2).value)
        owner = cell_text(sheet.cell(3, 4).value)
        cutoff = cell_text(sheet.cell(3, 5).value)
        claim_start = cell_text(sheet.cell(5, 3).value)
        claim_end = cell_text(sheet.cell(6, 3).value)

        master_codes.extend(find_master_codes(sheet, title, work_status, claim_end, item_map))

        header_row = find_header_row(sheet)
        if not header_row:
            continue

        participants = {}
        pending_rewards = []

        for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, values_only=True):
            values = [cell_text(value) for value in row]
            no = values[0] if len(values) > 0 else ""
            reward = values[1] if len(values) > 1 else ""
            facebook = values[2] if len(values) > 2 else ""
            uid = values[3] if len(values) > 3 else ""
            method = values[4] if len(values) > 4 else ""
            status = values[5] if len(values) > 5 else ""
            note = values[7] if len(values) > 7 else ""

            if not any([no, reward, facebook, uid]):
                continue
            if reward == "Reward" or no == "NO.":
                continue

            reward_info = build_reward_info(values, item_map) or {
                "name": reward or "รางวัลกิจกรรม",
                "forumReward": reward or "รางวัลกิจกรรม",
                "itemEn": "",
                "itemId": "",
                "itemType": "",
                "ownPeriod": "",
                "amount": "1",
                "imageUrl": "",
                "hasItem": False,
            }

            if facebook or uid:
                key = normalize_text(uid or facebook)
                entry = participants.setdefault(
                    key,
                    {
                        "uid": uid,
                        "facebook": facebook or "-",
                        "character": uid or "-",
                        "claimMethod": method,
                        "claimStatus": status or work_status or "กำลังดำเนินการ",
                        "updatedAt": delivery_date or claim_start or "",
                        "note": note,
                        "rewards": [],
                    },
                )
                if facebook and entry["facebook"] == "-":
                    entry["facebook"] = facebook
                if uid and not entry["uid"]:
                    entry["uid"] = uid
                if method and not entry["claimMethod"]:
                    entry["claimMethod"] = method
                if status:
                    entry["claimStatus"] = status
                if note:
                    entry["note"] = note
                reward_key = (
                    reward_info["itemId"]
                    or reward_info["imageUrl"]
                    or reward_info["itemEn"]
                    or reward_info["forumReward"]
                    or reward_info["name"]
                )
                if reward_key and reward_key not in {
                    item["itemId"] or item["imageUrl"] or item["itemEn"] or item["forumReward"] or item["name"]
                    for item in entry["rewards"]
                }:
                    entry["rewards"].append(reward_info)
            elif reward_info["hasItem"]:
                pending_rewards.append(reward_info)

        event_number = len(events) + 1
        events.append(
            {
                "id": f"event-{event_number}",
                "name": title,
                "shortName": title[:18],
                "icon": ICONS[(event_number - 1) % len(ICONS)],
                "cycle": f"กิจกรรมที่ {event_number}",
                "period": f"{claim_start or '-'} - {claim_end or '-'}",
                "resetDate": f"ตัดรอบ/จัดส่ง: {cutoff or 'ทุกวันพุธ'}",
                "latest": delivery_date or "รออัปเดต",
                "status": work_status or "กำลังดำเนินการ",
                "owner": owner,
                "reward": ", ".join(item["itemEn"] or item["name"] for item in pending_rewards[:3])
                or "ดูรางวัลในรายชื่อผู้ได้รับรางวัล",
                "winners": sorted(participants.values(), key=lambda w: bool(w.get("note"))),
                "pendingRewards": pending_rewards[:20],
            }
        )

    return events, master_codes


def main():
    source = resolve_source()
    workbook = openpyxl.load_workbook(source, data_only=True)
    events, master_codes = build_events(workbook)

    payload = (
        "window.WARZ_EVENTS = "
        + json.dumps(events, ensure_ascii=False, indent=2)
        + ";\n\nwindow.WARZ_MASTER_CODES = "
        + json.dumps(master_codes, ensure_ascii=False, indent=2)
        + ";\n"
    )
    OUTPUT.write_text(payload, encoding="utf-8")

    print(
        json.dumps(
            {
                "source": str(source),
                "events": len(events),
                "masterCodes": len(master_codes),
                "eventCounts": [{"name": event["name"], "winners": len(event["winners"])} for event in events],
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
